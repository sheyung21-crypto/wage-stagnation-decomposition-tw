from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .crosswalk import is_major_label, normalize_label
from .provenance import sha256_file


MAJOR_INDUSTRIES = [
    ("B", "mining", 8),
    ("C", "manufacturing", 11),
    ("D", "electricity_gas", 14),
    ("E", "water_waste", 18),
    ("F", "construction", 21),
    ("G", "wholesale_retail", 27),
    ("H", "transportation_storage", 30),
    ("I", "accommodation_food", 34),
    ("J", "information_communication", 37),
    ("K", "finance_insurance", 40),
    ("L", "real_estate", 43),
    ("M", "professional_scientific", 46),
    ("N", "support_services", 50),
    ("P", "education", 53),
    ("Q", "health_social", 56),
    ("R", "arts_recreation", 59),
    ("S", "other_services", 62),
]


def _number(value: object) -> float | None:
    if value is None or str(value).strip() in {"", "-", "…", "..."}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _year(value: object) -> int | None:
    match = re.search(r"(?:19|20)\d{2}", str(value or ""))
    return int(match.group()) if match else None


def _read_major_sheet(path: Path, sheet_name: str | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    records: list[dict[str, object]] = []
    totals: list[dict[str, float | int]] = []
    for row in worksheet.iter_rows(values_only=True):
        year = _year(row[0] if row else None)
        official = _number(row[1] if len(row) > 1 else None)
        if year is None or official is None:
            continue
        totals.append({"year": year, "official": official})
        for code, industry, column in MAJOR_INDUSTRIES:
            value = _number(row[column - 1] if len(row) >= column else None)
            if value is None:
                continue
            records.append({"year": year, "industry_code": code, "industry": industry, "value": value})
    workbook.close()
    return pd.DataFrame(records), pd.DataFrame(totals).drop_duplicates("year")


def _read_major_release(raw_dir: Path, suffix: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    paths = {
        "employees": raw_dir / f"s3_annual_employees_by_industry{suffix}.xlsx",
        "total_monthly": raw_dir / f"s2_annual_total_wage_by_industry{suffix}.xlsx",
        "regular_monthly": raw_dir / f"s1_annual_regular_wage_by_industry{suffix}.xlsx",
        "total_hours": raw_dir / f"s4_annual_hours_by_industry{suffix}.xlsx",
        "regular_hours": raw_dir / f"s4_annual_hours_by_industry{suffix}.xlsx",
    }
    sheets = {"total_hours": "t36-1", "regular_hours": "t36-2"}
    frames: list[pd.DataFrame] = []
    official_frames: list[pd.DataFrame] = []
    for metric, path in paths.items():
        frame, official = _read_major_sheet(path, sheets.get(metric))
        frames.append(frame.rename(columns={"value": metric}))
        official_frames.append(official.rename(columns={"official": f"official_{metric}"}))
    panel = frames[0]
    for frame in frames[1:]:
        panel = panel.merge(frame, on=["year", "industry_code", "industry"], validate="one_to_one")
    official = official_frames[0]
    for frame in official_frames[1:]:
        official = official.merge(frame, on="year", validate="one_to_one")
    return panel, official


def build_major_panel(raw_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    legacy, legacy_official = _read_major_release(raw_dir, "_2019_release")
    current, current_official = _read_major_release(raw_dir, "")
    legacy = legacy.loc[legacy.year.between(2000, 2004)].assign(source_release=2019, tsic_version=10)
    current = current.loc[current.year.between(2005, 2024)].assign(source_release=2024, tsic_version=11)
    legacy_official = legacy_official.loc[legacy_official.year.between(2000, 2004)].assign(source_release=2019)
    current_official = current_official.loc[current_official.year.between(2005, 2024)].assign(source_release=2024)
    panel = pd.concat([legacy, current], ignore_index=True).sort_values(["year", "industry_code"])
    official = pd.concat([legacy_official, current_official], ignore_index=True).sort_values("year")
    counts = panel.groupby("year").size()
    expected = pd.Series({year: 16 if year <= 2008 else 17 for year in range(2000, 2025)})
    if not counts.equals(expected.rename(None)):
        raise ValueError(f"Major-industry panel has unexpected coverage: {counts.to_dict()}")
    return panel.reset_index(drop=True), official.reset_index(drop=True)


def build_vintage_major_panel(raw_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    panels: list[pd.DataFrame] = []
    officials: list[pd.DataFrame] = []
    for roc_year in range(105, 113):
        year = roc_year + 1911
        paths = {
            "employees": raw_dir / f"{roc_year}_major_table01.xlsx",
            "total_monthly": raw_dir / f"{roc_year}_major_table18.xlsx",
            "regular_monthly": raw_dir / f"{roc_year}_major_table19.xlsx",
            "total_hours": raw_dir / f"{roc_year}_major_table36.xlsx",
            "regular_hours": raw_dir / f"{roc_year}_major_table36.xlsx",
        }
        frames: list[pd.DataFrame] = []
        official_frames: list[pd.DataFrame] = []
        for metric, path in paths.items():
            frame, official = _read_major_sheet(path, {"total_hours": "t36-1", "regular_hours": "t36-2"}.get(metric))
            frames.append(frame.loc[frame.year.eq(year)].rename(columns={"value": metric}))
            official_frames.append(official.loc[official.year.eq(year)].rename(columns={"official": f"official_{metric}"}))
        panel = frames[0]
        for frame in frames[1:]:
            panel = panel.merge(frame, on=["year", "industry_code", "industry"], validate="one_to_one")
        official = official_frames[0]
        for frame in official_frames[1:]:
            official = official.merge(frame, on="year", validate="one_to_one")
        panels.append(panel.assign(source_release=year, tsic_version=10 if year <= 2020 else 11))
        officials.append(official.assign(source_release=year))
    current, current_official = _read_major_release(raw_dir, "")
    panels.append(current.loc[current.year.eq(2024)].assign(source_release=2024, tsic_version=11))
    officials.append(current_official.loc[current_official.year.eq(2024)].assign(source_release=2024))
    return pd.concat(panels, ignore_index=True), pd.concat(officials, ignore_index=True)


def _header_column(worksheet, token: str) -> int:
    for row in worksheet.iter_rows(min_row=1, max_row=min(12, worksheet.max_row)):
        for cell in row:
            if token in str(cell.value or ""):
                return cell.column
    raise ValueError(f"Could not locate column headed by {token}")


def _read_middle_metric(path: Path, kind: str) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, object]] = []
    skipped_major_labels: set[str] = set()
    for worksheet in workbook.worksheets:
        if "old" in worksheet.title.lower():
            continue
        secondary_column = None
        if kind == "wage":
            secondary_column = _header_column(worksheet, "經常性")
        elif kind == "hours":
            secondary_column = _header_column(worksheet, "正常")
        for row in worksheet.iter_rows(values_only=True):
            label = row[0] if row else None
            total = _number(row[1] if len(row) > 1 else None)
            label_text = str(label or "")
            if (
                not label
                or total is None
                or "\n" not in label_text
                or not re.search(r"[A-Za-z]", label_text.split("\n", 1)[1])
            ):
                continue
            normalized = normalize_label(label)
            if is_major_label(label) and normalized not in skipped_major_labels:
                skipped_major_labels.add(normalized)
                continue
            record: dict[str, object] = {"industry": normalized, "industry_name": str(label).split("\n", 1)[0].strip()}
            if kind == "employment":
                record["employees"] = total
            elif kind == "wage":
                regular = _number(row[secondary_column - 1] if len(row) >= secondary_column else None)
                if regular is None:
                    continue
                record.update(total_monthly=total, regular_monthly=regular)
            else:
                regular = _number(row[secondary_column - 1] if len(row) >= secondary_column else None)
                if regular is None:
                    continue
                record.update(total_hours=total, regular_hours=regular)
            records.append(record)
    workbook.close()
    frame = pd.DataFrame(records)
    duplicates_removed = 0
    if not frame.empty and frame.industry.duplicated().any():
        duplicated = frame.loc[frame.industry.duplicated(keep=False)]
        value_columns = [column for column in frame.columns if column not in {"industry", "industry_name"}]
        if duplicated.groupby("industry")[value_columns].nunique(dropna=False).to_numpy().max() > 1:
            raise ValueError(f"Conflicting duplicate middle-industry rows in {path.name}")
        duplicates_removed = int(frame.industry.duplicated().sum())
        frame = frame.drop_duplicates("industry", keep="first")
    if frame.empty:
        raise ValueError(f"No middle-industry rows in {path.name}")
    frame.attrs["duplicate_rows_removed"] = duplicates_removed
    return frame


def build_middle_panel(raw_dir: Path) -> pd.DataFrame:
    all_records: list[pd.DataFrame] = []
    exclusions: list[dict[str, object]] = []
    for year in range(2016, 2025):
        roc_year = year - 1911
        for offset in range(14):
            employment = _read_middle_metric(raw_dir / f"{roc_year}_table{4 + offset:02d}.xlsx", "employment")
            wage = _read_middle_metric(raw_dir / f"{roc_year}_table{22 + offset:02d}.xlsx", "wage")
            hours = _read_middle_metric(raw_dir / f"{roc_year}_table{39 + offset:02d}.xlsx", "hours")
            source_duplicates = sum(frame.attrs.get("duplicate_rows_removed", 0) for frame in (employment, wage, hours))
            if source_duplicates:
                exclusions.append(
                    {
                        "year": year,
                        "table_group": offset + 1,
                        "employment_extra": "official duplicate rows",
                        "wage_extra": "",
                        "hours_extra": "",
                        "rows_excluded": source_duplicates,
                    }
                )
            key = "industry"
            name_sets = [set(frame[key]) for frame in (employment, wage, hours)]
            common = set.intersection(*name_sets)
            if not common:
                raise ValueError(f"No common middle-industry labels in {year}, table group {offset + 1}")
            extras = [sorted(names - common) for names in name_sets]
            if any(extras):
                exclusions.append(
                    {
                        "year": year,
                        "table_group": offset + 1,
                        "employment_extra": "|".join(extras[0]),
                        "wage_extra": "|".join(extras[1]),
                        "hours_extra": "|".join(extras[2]),
                        "rows_excluded": sum(map(len, extras)),
                    }
                )
            employment = employment.loc[employment.industry.isin(common)]
            wage = wage.loc[wage.industry.isin(common)]
            hours = hours.loc[hours.industry.isin(common)]
            merged = employment.merge(wage.drop(columns="industry_name"), on=key, validate="one_to_one")
            merged = merged.merge(hours.drop(columns="industry_name"), on=key, validate="one_to_one")
            merged["year"] = year
            merged["table_group"] = offset + 1
            merged["tsic_version"] = 10 if year <= 2020 else 11
            merged["regime"] = "tsic10_2016_2020" if year <= 2020 else "tsic11_2021_2024"
            all_records.append(merged)
    panel = pd.concat(all_records, ignore_index=True)
    if panel.duplicated(["year", "industry"]).any():
        numeric = ["employees", "total_monthly", "regular_monthly", "total_hours", "regular_hours"]
        duplicated = panel.loc[panel.duplicated(["year", "industry"], keep=False)]
        spread = duplicated.groupby(["year", "industry"])[numeric].agg(lambda values: float(values.max() - values.min()))
        if (spread.to_numpy() > 1e-10).any():
            raise ValueError("Duplicated adjacent-table rows disagree numerically")
        duplicate_count = int(panel.duplicated(["year", "industry"]).sum())
        exclusions.append(
            {
                "year": "2016-2024",
                "table_group": "combined_workbooks",
                "employment_extra": "",
                "wage_extra": "",
                "hours_extra": "",
                "rows_excluded": duplicate_count,
            }
        )
        panel = panel.drop_duplicates(["year", "industry"], keep="first")
    panel = panel.sort_values(["year", "table_group", "industry"]).reset_index(drop=True)
    panel.attrs["source_exclusions"] = exclusions
    return panel


def read_annual_cpi(path: Path, start: int = 2000, end: int = 2024) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for _, element in ET.iterparse(path, events=("end",)):
        if element.tag != "Obs":
            continue
        values = {child.tag: child.text for child in element}
        item = values.get("Item") or ""
        period = values.get("TIME_PERIOD") or ""
        if item.startswith("總指數") and values.get("FREQ") == "M" and values.get("TYPE") == "原始值":
            match = re.fullmatch(r"(\d{4})M(\d{2})", period)
            value = _number(values.get("Item_VALUE"))
            if match and value is not None:
                year = int(match.group(1))
                if start <= year <= end:
                    rows.append({"year": year, "month": int(match.group(2)), "cpi": value})
        element.clear()
    monthly = pd.DataFrame(rows)
    counts = monthly.groupby("year").month.nunique()
    if len(counts) != end - start + 1 or counts.ne(12).any():
        raise ValueError("CPI input does not contain twelve months for every analysis year")
    return monthly.groupby("year", as_index=False).cpi.mean()


def read_official_real_wage(path: Path, start: int = 2000, end: int = 2024) -> pd.DataFrame:
    """Read the annual real-wage sequence exported from the official query system."""
    frame = pd.read_csv(path)
    expected_columns = {
        "year",
        "official_real_total_monthly",
        "official_real_regular_monthly",
    }
    if set(frame.columns) != expected_columns:
        raise ValueError("Official real-wage extract has unexpected columns")
    expected_years = list(range(start, end + 1))
    if frame.year.tolist() != expected_years:
        raise ValueError("Official real-wage extract does not cover every requested year")
    if (frame.drop(columns="year") <= 0).any().any():
        raise ValueError("Official real-wage values must be positive")
    return frame


def expected_raw_sources() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    current_base = "https://ws.dgbas.gov.tw/001/Upload/463/relfile/11049/234856"
    legacy_base = "https://ws.dgbas.gov.tw/public/data/dgbas04/bc5/month/10813"
    major = [
        ("s3_annual_employees_by_industry.xlsx", f"{current_base}/table1.xlsx", "2005-2024", "11"),
        ("s2_annual_total_wage_by_industry.xlsx", f"{current_base}/table18.xlsx", "2005-2024", "11"),
        ("s1_annual_regular_wage_by_industry.xlsx", f"{current_base}/table19.xlsx", "2005-2024", "11"),
        ("s4_annual_hours_by_industry.xlsx", f"{current_base}/table36.xlsx", "2005-2024", "11"),
        ("s3_annual_employees_by_industry_2019_release.xlsx", f"{legacy_base}/table1.xlsx", "2000-2019", "10"),
        ("s2_annual_total_wage_by_industry_2019_release.xlsx", f"{legacy_base}/table18.xlsx", "2000-2019", "10"),
        ("s1_annual_regular_wage_by_industry_2019_release.xlsx", f"{legacy_base}/table19.xlsx", "2000-2019", "10"),
        ("s4_annual_hours_by_industry_2019_release.xlsx", f"{legacy_base}/table36.xlsx", "2000-2019", "10"),
    ]
    for filename, url, coverage, version in major:
        rows.append(dict(filename=filename, url=url, coverage=coverage, classification_version=version, notes="major-industry annual table"))
    release_ids = {111: "231487", 112: "233568", 113: "234856"}
    for roc_year in range(105, 114):
        if roc_year <= 110:
            base = f"https://ws.dgbas.gov.tw/public/data/dgbas04/bc5/month/{roc_year}13"
        else:
            base = f"https://ws.dgbas.gov.tw/001/Upload/463/relfile/11049/{release_ids[roc_year]}"
        for table in list(range(4, 18)) + list(range(22, 36)) + list(range(39, 53)):
            rows.append(
                dict(
                    filename=f"{roc_year}_table{table:02d}.xlsx",
                    url=f"{base}/table{table}.xlsx",
                    coverage=str(roc_year + 1911),
                    classification_version="10" if roc_year <= 109 else "11",
                    notes="middle-industry annual table",
                )
            )
        if roc_year <= 112:
            for table in (1, 18, 19, 36):
                rows.append(
                    dict(
                        filename=f"{roc_year}_major_table{table:02d}.xlsx",
                        url=f"{base}/table{table}.xlsx",
                        coverage=f"historical through {roc_year + 1911}; analysis uses {roc_year + 1911}",
                        classification_version="10" if roc_year <= 109 else "11",
                        notes="same-vintage major-industry validation table",
                    )
                )
    rows.extend(
        [
            dict(filename="official_real_wage_series.csv", url="https://earnings.dgbas.gov.tw/query_payroll.aspx", coverage="2000-2024", classification_version="official_published_2024_vintage", notes="official published annual real regular and total earnings query"),
            dict(filename="cpi_basic_classification.xml", url="https://ws.dgbas.gov.tw/001/Upload/461/relfile/11525/230555/pr0101a1m.xml", coverage="1981-latest", classification_version="not_applicable", notes="monthly all-items CPI"),
            dict(filename="dgbas_open_data_108.xml", url="https://ws.dgbas.gov.tw/001/Upload/461/relfile/11525/233528/Mp07014A108.xml", coverage="2019 release index", classification_version="10", notes="official open-data source index"),
            dict(filename="tsic_7_to_8_crosswalk.pdf", url="https://ws.dgbas.gov.tw/001/Upload/463/relfile/11195/90002/80f2bef2-0296-4ec3-896f-6490f2fa3f41.pdf", coverage="TSIC 7 to 8", classification_version="7-8", notes="official crosswalk"),
            dict(filename="tsic_8_to_9_crosswalk.xls", url="https://ws.dgbas.gov.tw/001/Upload/463/attachment/11195/90008/18119504471.xls", coverage="TSIC 8 to 9", classification_version="8-9", notes="official crosswalk"),
            dict(filename="tsic_9_to_10_crosswalk.odt", url="https://ws.dgbas.gov.tw/001/Upload/463/attachment/11195/90010/6112105911geyjeag4.odt", coverage="TSIC 9 to 10", classification_version="9-10", notes="official crosswalk"),
            dict(filename="tsic_10_to_11_crosswalk.pdf", url="https://ws.dgbas.gov.tw/001/Upload/463/relfile/11195/90015/%E6%9E%B6%E6%A7%8B_10%E5%B0%8D11_1141205%E6%9B%B4%E6%96%B0.pdf", coverage="TSIC 10 to 11", classification_version="10-11", notes="official crosswalk"),
            dict(filename="mol_foreign_workers_by_work.csv", url="https://statdb.mol.gov.tw/html/year/year13/313020.htm", coverage="1996-2024; analysis uses 2000-2024", classification_version="MOL work-responsibility categories; TSIC revisions not used", notes="MOL year-end migrant-worker stock by work responsibility; 2015 onward industry split uses valid permits"),
            dict(filename="mol_major_economic_indicators.csv", url="https://statdb.mol.gov.tw/html/year/year14/31010.htm", coverage="1996-2025; analysis uses 2000-2024", classification_version="not_applicable", notes="MOL major economic indicators: year-end monthly minimum wage and manufacturing productivity index"),
            dict(filename="dgbas_manufacturing_labor_productivity.xlsx", url="https://www.dgbas.gov.tw/news.aspx?n=4437&sms=10980&_CSN=135", coverage="2000-2024", classification_version="official manufacturing series; 2021=100", notes="DGBAS manufacturing production-volume-per-hour productivity index used for dissemination cross-check"),
            dict(filename="dgbas_national_accounts_yearbook.pdf", url="https://ws.dgbas.gov.tw/001/Upload/463/relfile/11025/235627/nay.pdf", coverage="national accounts through 2024", classification_version="SNA 2008; TSIC 11", notes="DGBAS National Accounts Yearbook documentation and definitions"),
            dict(filename="dgbas_national_accounts_abstract.pdf", url="https://ebook.dgbas.gov.tw/News_eBook.aspx?_CSN=610&n=3786&sms=11503", coverage="national accounts through May 2026 release", classification_version="SNA 2008; TSIC 11", notes="DGBAS Statistical Abstract of National Accounts documentation"),
            dict(filename="dgbas_national_accounts_compensation_metadata.json", url="https://nstatdb.dgbas.gov.tw/dgbasAll/webMain.aspx?sdmx/A018102020/all", coverage="series metadata", classification_version="SNA 2008; TSIC 11", notes="DGBAS SDMX metadata for domestic production and factor income"),
            dict(filename="dgbas_national_accounts_compensation.json", url="https://nstatdb.dgbas.gov.tw/dgbasAll/webMain.aspx?sdmx/A018102020/3+6.6+7+33+36+39+40+43+49+52+56+60+63+64+67+68+69+72+73+77+80+84..A&startTime=2000-00&endTime=2024-00", coverage="2000-2024", classification_version="SNA 2008; TSIC 11", notes="DGBAS SDMX annual nominal value added and employee compensation by industry"),
            dict(filename="dgbas_national_accounts_output.json", url="https://nstatdb.dgbas.gov.tw/dgbasAll/webMain.aspx?sdmx/A018103010/3+4+5.6+7+33+36+39+40+43+49+52+56+60+63+64+67+68+69+72+73+77+80+82+84+85..A&startTime=2000-00&endTime=2024-00", coverage="2000-2024", classification_version="SNA 2008; TSIC 11", notes="DGBAS SDMX annual nominal output, chain-volume output, and output deflators by industry"),
        ]
    )
    return rows


def build_source_manifest(raw_dir: Path, output: Path) -> pd.DataFrame:
    records = []
    for source in expected_raw_sources():
        path = raw_dir / str(source["filename"])
        if not path.exists():
            raise FileNotFoundError(path)
        records.append(
            {
                "file": f"data/raw/{path.name}",
                "url": source["url"],
                "acquired_at": "2026-08-09",
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
                "coverage": source["coverage"],
                "classification_version": source["classification_version"],
                "notes": source["notes"],
            }
        )
    frame = pd.DataFrame(records).sort_values("file")
    frame.to_csv(output, index=False, encoding="utf-8-sig", lineterminator="\n")
    return frame
